
import openpyxl

def inspect_excel(filename):
    print(f"--- Inspecting {filename} ---")
    wb = openpyxl.load_workbook(filename, data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"Sheet: {sheet_name}")
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.replace('\n', ' ')
                    # Print coordinate and value, manually handling basic Japanese strings if possible
                    # or just hex-encode if it fails, but Python 3 usually handles UTF-8 print.
                    # The terminal might be the issue.
                    try:
                        print(f"{cell.coordinate}: {val}")
                    except:
                        print(f"{cell.coordinate}: [ENCODING ERROR]")

inspect_excel('c:\\Users\\user\\Desktop\\syako\\2-2syoumeisinsei0803.xlsx')
inspect_excel('c:\\Users\\user\\Desktop\\syako\\3-2hokanbasyotodoke0803.xlsx')
