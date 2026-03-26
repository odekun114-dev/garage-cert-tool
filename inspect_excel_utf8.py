
import openpyxl
import io

def inspect_excel(filename):
    out = io.StringIO()
    out.write(f"--- Inspecting {filename} ---\n")
    wb = openpyxl.load_workbook(filename, data_only=True)
    for sheet_name in wb.sheetnames:
        out.write(f"Sheet: {sheet_name}\n")
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    val = str(cell.value).replace('\n', ' ')
                    out.write(f"{cell.coordinate}: {val}\n")
    return out.getvalue()

with open('c:\\Users\\user\\Desktop\\syako\\excel_inspection_utf8.txt', 'w', encoding='utf-8') as f:
    f.write(inspect_excel('c:\\Users\\user\\Desktop\\syako\\2-2syoumeisinsei0803.xlsx'))
    f.write(inspect_excel('c:\\Users\\user\\Desktop\\syako\\3-2hokanbasyotodoke0803.xlsx'))
